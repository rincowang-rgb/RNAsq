"""
RNA-seq Analysis Module
Analyzes RNA-seq data from Excel files with support for control vs experimental comparison,
statistical filtering, and literature preparation.
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import numpy as np
from typing import Tuple, Dict, List, Optional


class RNASeqAnalyzer:
    """Main class for RNA-seq analysis operations."""
    
    def __init__(self, excel_file: str):
        """
        Initialize the RNA-seq analyzer with an Excel file.
        
        Args:
            excel_file: Path to the Excel file containing RNA-seq data
        """
        self.excel_file = excel_file
        self.data = None
        self.workbook = None
        
    def read_excel_file(self, sheet_name: str = 0) -> pd.DataFrame:
        """
        Read RNA-seq data from Excel file.
        
        Args:
            sheet_name: Sheet name or index to read (default: 0)
            
        Returns:
            DataFrame containing the RNA-seq data
        """
        try:
            self.data = pd.read_excel(self.excel_file, sheet_name=sheet_name)
            print(f"Successfully loaded data from {self.excel_file}")
            print(f"Data shape: {self.data.shape}")
            return self.data
        except Exception as e:
            print(f"Error reading Excel file: {str(e)}")
            return None
    
    def load_workbook(self):
        """Load openpyxl workbook for advanced operations."""
        try:
            self.workbook = openpyxl.load_workbook(self.excel_file)
            print(f"Workbook loaded: {self.excel_file}")
            return self.workbook
        except Exception as e:
            print(f"Error loading workbook: {str(e)}")
            return None
    
    def compare_control_vs_experimental(
        self,
        control_columns: List[str],
        experimental_columns: List[str],
        gene_column: str = 'Gene_ID'
    ) -> pd.DataFrame:
        """
        Compare gene expression between control and experimental conditions.
        
        Args:
            control_columns: List of column names for control samples
            experimental_columns: List of column names for experimental samples
            gene_column: Name of the gene identifier column
            
        Returns:
            DataFrame with comparison results including fold change and p-values
        """
        if self.data is None:
            print("No data loaded. Call read_excel_file() first.")
            return None
        
        result = self.data.copy()
        
        # Calculate mean expression for each group
        result['control_mean'] = self.data[control_columns].mean(axis=1)
        result['experimental_mean'] = self.data[experimental_columns].mean(axis=1)
        
        # Calculate fold change (avoid division by zero)
        result['fold_change'] = np.where(
            result['control_mean'] > 0,
            result['experimental_mean'] / result['control_mean'],
            np.nan
        )
        
        # Calculate log2 fold change
        result['log2_fold_change'] = np.log2(result['fold_change'].replace(0, np.nan))
        
        # Calculate standard deviation for each group
        result['control_std'] = self.data[control_columns].std(axis=1)
        result['experimental_std'] = self.data[experimental_columns].std(axis=1)
        
        # Perform t-test (simplified p-value calculation)
        from scipy.stats import ttest_ind
        
        p_values = []
        for idx, row in self.data.iterrows():
            try:
                control_vals = self.data.loc[idx, control_columns].values
                exp_vals = self.data.loc[idx, experimental_columns].values
                _, p_val = ttest_ind(control_vals, exp_vals)
                p_values.append(p_val)
            except:
                p_values.append(np.nan)
        
        result['p_value'] = p_values
        result['neg_log10_pvalue'] = -np.log10(result['p_value'])
        
        print(f"Comparison complete: {len(result)} genes analyzed")
        return result
    
    def filter_significant_genes(
        self,
        comparison_df: pd.DataFrame,
        fold_change_threshold: float = 2.0,
        p_value_threshold: float = 0.05,
        log2_fc_column: str = 'log2_fold_change',
        p_value_column: str = 'p_value'
    ) -> pd.DataFrame:
        """
        Filter genes based on fold change and p-value thresholds.
        
        Args:
            comparison_df: DataFrame from compare_control_vs_experimental()
            fold_change_threshold: Minimum absolute fold change threshold
            p_value_threshold: Maximum p-value threshold
            log2_fc_column: Name of log2 fold change column
            p_value_column: Name of p-value column
            
        Returns:
            Filtered DataFrame containing only significant genes
        """
        if comparison_df is None:
            print("No comparison data provided.")
            return None
        
        # Filter based on thresholds
        filtered = comparison_df[
            (abs(comparison_df[log2_fc_column]) >= np.log2(fold_change_threshold)) &
            (comparison_df[p_value_column] < p_value_threshold)
        ].copy()
        
        # Classify as upregulated or downregulated
        filtered['regulation'] = filtered[log2_fc_column].apply(
            lambda x: 'upregulated' if x > 0 else 'downregulated'
        )
        
        # Sort by log2 fold change
        filtered = filtered.sort_values(by=log2_fc_column, ascending=False)
        
        print(f"Filtered genes: {len(filtered)}")
        print(f"  Upregulated: {len(filtered[filtered['regulation'] == 'upregulated'])}")
        print(f"  Downregulated: {len(filtered[filtered['regulation'] == 'downregulated'])}")
        
        return filtered
    
    def prepare_literature_comparison(
        self,
        filtered_df: pd.DataFrame,
        include_columns: Optional[List[str]] = None
    ) -> pd.DataFrame:
        """
        Prepare filtered data for literature comparison and export.
        
        Args:
            filtered_df: DataFrame from filter_significant_genes()
            include_columns: Specific columns to include (optional)
            
        Returns:
            DataFrame formatted for literature comparison
        """
        if filtered_df is None or filtered_df.empty:
            print("No filtered data to prepare.")
            return None
        
        # Default columns to include
        default_columns = [
            'Gene_ID', 'log2_fold_change', 'p_value', 'neg_log10_pvalue',
            'regulation', 'control_mean', 'experimental_mean'
        ]
        
        # Use provided columns or defaults
        cols_to_keep = include_columns if include_columns else default_columns
        cols_to_keep = [col for col in cols_to_keep if col in filtered_df.columns]
        
        literature_df = filtered_df[cols_to_keep].copy()
        
        # Add literature-ready formatting
        literature_df['abs_log2_fc'] = abs(literature_df['log2_fold_change'])
        literature_df = literature_df.sort_values(by='abs_log2_fc', ascending=False)
        
        # Round numerical values for readability
        float_cols = literature_df.select_dtypes(include=['float64', 'float32']).columns
        for col in float_cols:
            literature_df[col] = literature_df[col].round(4)
        
        print(f"Literature comparison data prepared: {len(literature_df)} genes")
        return literature_df
    
    def export_results(
        self,
        data: pd.DataFrame,
        output_file: str,
        sheet_name: str = 'Results'
    ) -> bool:
        """
        Export analysis results to Excel file.
        
        Args:
            data: DataFrame to export
            output_file: Path to output Excel file
            sheet_name: Name of the sheet to create
            
        Returns:
            True if successful, False otherwise
        """
        try:
            data.to_excel(output_file, sheet_name=sheet_name, index=False)
            print(f"Results exported to {output_file}")
            return True
        except Exception as e:
            print(f"Error exporting results: {str(e)}")
            return False
    
    def get_summary_statistics(self, data: pd.DataFrame) -> Dict:
        """
        Generate summary statistics for the dataset.
        
        Args:
            data: DataFrame to analyze
            
        Returns:
            Dictionary containing summary statistics
        """
        if data is None or data.empty:
            return {}
        
        stats = {
            'total_genes': len(data),
            'upregulated': len(data[data.get('regulation') == 'upregulated']),
            'downregulated': len(data[data.get('regulation') == 'downregulated']),
            'mean_log2_fc': data.get('log2_fold_change', pd.Series()).mean(),
            'median_log2_fc': data.get('log2_fold_change', pd.Series()).median(),
            'max_log2_fc': data.get('log2_fold_change', pd.Series()).max(),
            'min_log2_fc': data.get('log2_fold_change', pd.Series()).min(),
        }
        
        return stats


def analyze_rna_seq_file(
    excel_file: str,
    control_cols: List[str],
    experimental_cols: List[str],
    gene_col: str = 'Gene_ID',
    fc_threshold: float = 2.0,
    pval_threshold: float = 0.05,
    output_file: Optional[str] = None
) -> Tuple[pd.DataFrame, Dict]:
    """
    Complete RNA-seq analysis pipeline.
    
    Args:
        excel_file: Path to input Excel file
        control_cols: List of control sample column names
        experimental_cols: List of experimental sample column names
        gene_col: Gene identifier column name
        fc_threshold: Fold change threshold
        pval_threshold: P-value threshold
        output_file: Optional output file path
        
    Returns:
        Tuple of (filtered DataFrame, summary statistics)
    """
    # Initialize analyzer
    analyzer = RNASeqAnalyzer(excel_file)
    
    # Read data
    analyzer.read_excel_file()
    
    # Compare conditions
    comparison = analyzer.compare_control_vs_experimental(
        control_columns=control_cols,
        experimental_columns=experimental_cols,
        gene_column=gene_col
    )
    
    # Filter significant genes
    filtered = analyzer.filter_significant_genes(
        comparison,
        fold_change_threshold=fc_threshold,
        p_value_threshold=pval_threshold
    )
    
    # Prepare for literature
    literature_ready = analyzer.prepare_literature_comparison(filtered)
    
    # Get statistics
    stats = analyzer.get_summary_statistics(filtered)
    
    # Export if requested
    if output_file:
        analyzer.export_results(literature_ready, output_file)
    
    return literature_ready, stats


if __name__ == "__main__":
    # Example usage
    print("RNA-seq Analysis Module")
    print("Usage: from rna_seq_analysis import RNASeqAnalyzer, analyze_rna_seq_file")
